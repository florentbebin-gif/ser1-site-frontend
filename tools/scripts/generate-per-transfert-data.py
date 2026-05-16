"""
Genere les donnees PER transfert depuis les sources de reference locales/publiques.

Sources :
- Classeur Excel local, onglet BASECG.
- CASdatasets, fichiers freTGF05/freTGH05/freTPRV93/freTPG93full.

Prerequis Python :
    pip install openpyxl pyreadr

Usage :
    python tools/scripts/generate-per-transfert-data.py
"""

from __future__ import annotations

import json
import os
import re
import sys
import urllib.request
from pathlib import Path
from typing import Any

try:
    import openpyxl
except ImportError as exc:  # pragma: no cover - message operateur
    raise SystemExit("openpyxl est requis pour extraire BASECG.") from exc

try:
    import pyreadr
except ImportError as exc:  # pragma: no cover - message operateur
    raise SystemExit("pyreadr est requis pour lire les tables CASdatasets.") from exc


ROOT = Path(__file__).resolve().parents[2]
BASECG_OUT = ROOT / "src" / "data" / "basecg" / "catalog.generated.ts"
MORTALITY_DIR = ROOT / "src" / "data" / "mortality"
TMP_DIR = ROOT / ".tmp" / "per-transfert-data"

CASDATASETS_RAW = "https://raw.githubusercontent.com/dutangc/CASdatasets/master/data"
MORTALITY_FILES = {
    "tprv93": ("freTPRV93.rda", "freTPRV93"),
    "tgf05": ("freTGF05.rda", "freTGF05"),
    "tgh05": ("freTGH05.rda", "freTGH05"),
    "tpg93": ("freTPG93full.rda", "freTPG93full"),
}

BASECG_ROWS = {
    "sourceId": 1,
    "compagnie": 4,
    "nomContrat": 5,
    "versementAnnuel": 6,
    "dateEffet": 7,
    "dateCommercialisation": 8,
    "provisionMathematique": 9,
    "interets": 10,
    "nombreFonds": 11,
    "repartitionUcEuro": 12,
    "rendementFondsEuro": 13,
    "fraisVersements": 14,
    "fraisGestion": 15,
    "fraisArbitrage": 16,
    "fraisTransfertSortant": 17,
    "clauseBeneficiaire": 18,
    "garantiesComplementaires": 19,
    "ageLimiteLiquidation": 20,
    "sortieCapitalRetraite": 21,
    "fractionnementCapital": 22,
    "rachatLibre": 23,
    "tableConversionRente": 24,
    "tableGarantieAdhesion": 25,
    "tauxTechnique": 26,
    "fraisArrerages": 27,
    "annuitesGaranties": 28,
    "reversionPossible": 29,
    "reversionIncluse": 30,
    "renteEstimee": 31,
}


def resolve_xlsm_path() -> Path:
    explicit = os.environ.get("BASECG_XLSM_PATH")
    if explicit:
        return Path(explicit).expanduser().resolve()

    candidates = sorted(ROOT.glob("*.xlsm"))
    if len(candidates) != 1:
        raise SystemExit("Definir BASECG_XLSM_PATH vers le classeur BASECG.")
    return candidates[0]


XLSM_PATH = resolve_xlsm_path()


def slugify(value: str) -> str:
    lowered = value.lower()
    replacements = {
        "à": "a",
        "â": "a",
        "ä": "a",
        "ç": "c",
        "é": "e",
        "è": "e",
        "ê": "e",
        "ë": "e",
        "î": "i",
        "ï": "i",
        "ô": "o",
        "ö": "o",
        "ù": "u",
        "û": "u",
        "ü": "u",
        "ÿ": "y",
        "œ": "oe",
        "æ": "ae",
    }
    for src, dst in replacements.items():
        lowered = lowered.replace(src, dst)
    return re.sub(r"[^a-z0-9]+", "-", lowered).strip("-") or "contrat"


def ts_json(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, separators=(",", ":"))


def normalize_cell(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        cleaned = re.sub(r"[ \t]+", " ", value.replace("\r\n", "\n").replace("\r", "\n")).strip()
        return cleaned or None
    if hasattr(value, "isoformat"):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def detect_contract_type(company: str, name: str, table_label: str | None) -> str:
    haystack = f"{company} {name} {table_label or ''}".upper()
    if "PREFON" in haystack or "PRÉFON" in haystack:
        return "PER_POINTS"
    if "COREM" in haystack and "SYSTEME PAR POINTS" in haystack:
        return "PER_POINTS"
    if "ART83" in haystack or "ARTICLE 83" in haystack:
        return "ARTICLE83"
    if "MADELIN" in haystack:
        return "MADELIN"
    if "PERP" in haystack:
        return "PERP"
    if "PERCO" in haystack:
        return "PERCO"
    if "PERIN" in haystack or "PER INDIVIDUEL" in haystack or re.search(r"\bPER\b", haystack):
        return "PERIN"
    return "AUTRE"


def detect_compartment(contract_type: str, name: str) -> str:
    haystack = name.upper()
    if "NON DEDUIT" in haystack or "NON DÉDUIT" in haystack:
        return "C1_BIS"
    if contract_type == "ARTICLE83":
        return "C3"
    if contract_type == "PERCO":
        return "C2"
    return "C1"


def extract_rate(value: Any) -> float | None:
    if isinstance(value, (int, float)):
        number = float(value)
        return number if number <= 1 else number / 100
    if not isinstance(value, str):
        return None
    match = re.search(r"(\d+(?:[,.]\d+)?)\s*%", value)
    if not match:
        return None
    return float(match.group(1).replace(",", ".")) / 100


def extract_basecg() -> list[dict[str, Any]]:
    if not XLSM_PATH.exists():
        raise SystemExit(f"Classeur introuvable : {XLSM_PATH}")

    workbook = openpyxl.load_workbook(XLSM_PATH, data_only=True, read_only=False)
    sheet = workbook["BASECG"]
    contracts: list[dict[str, Any]] = []

    for col in range(3, sheet.max_column + 1):
        raw = {key: normalize_cell(sheet.cell(row=row, column=col).value) for key, row in BASECG_ROWS.items()}
        if not raw["compagnie"] or not raw["nomContrat"]:
            continue

        source_id = str(raw["sourceId"] or f"Contrat N°{col - 2}")
        company = str(raw["compagnie"])
        name = str(raw["nomContrat"])
        contract_type = detect_contract_type(company, name, raw.get("tableConversionRente"))
        contract = {
            "id": f"{slugify(company)}-{slugify(name)}-{col - 2}",
            "sourceId": source_id,
            "compagnie": company,
            "nomContrat": name,
            "typeContrat": contract_type,
            "perCompartment": detect_compartment(contract_type, name),
            "phaseEpargne": {
                "dateCommercialisation": raw["dateCommercialisation"],
                "nombreFonds": raw["nombreFonds"],
                "repartitionUcEuro": raw["repartitionUcEuro"],
                "rendementFondsEuro": raw["rendementFondsEuro"],
                "fraisVersements": raw["fraisVersements"],
                "fraisGestion": raw["fraisGestion"],
                "fraisArbitrage": raw["fraisArbitrage"],
                "fraisTransfertSortant": raw["fraisTransfertSortant"],
                "fraisTransfertSortantRate": extract_rate(raw["fraisTransfertSortant"]),
                "clauseBeneficiaire": raw["clauseBeneficiaire"],
                "garantiesComplementaires": raw["garantiesComplementaires"],
            },
            "phaseLiquidation": {
                "ageLimiteLiquidation": raw["ageLimiteLiquidation"],
                "sortieCapitalRetraite": raw["sortieCapitalRetraite"],
                "fractionnementCapital": raw["fractionnementCapital"],
                "rachatLibre": raw["rachatLibre"],
                "tableConversionRente": raw["tableConversionRente"],
                "tableGarantieAdhesion": raw["tableGarantieAdhesion"],
                "tauxTechnique": raw["tauxTechnique"],
                "fraisArrerages": raw["fraisArrerages"],
                "fraisArreragesRate": extract_rate(raw["fraisArrerages"]),
                "annuitesGaranties": raw["annuitesGaranties"],
                "reversionPossible": raw["reversionPossible"],
                "reversionIncluse": raw["reversionIncluse"],
                "renteEstimee": raw["renteEstimee"],
            },
        }
        contracts.append(contract)

    if len(contracts) < 390:
        raise SystemExit(f"Extraction BASECG incomplete : {len(contracts)} contrats.")

    return contracts


def write_basecg(contracts: list[dict[str, Any]]) -> None:
    BASECG_OUT.parent.mkdir(parents=True, exist_ok=True)
    lines = [
        "/* Fichier genere par tools/scripts/generate-per-transfert-data.py. */",
        "/* Source : classeur Excel local, onglet BASECG. */",
        "import type { BaseCgRetraiteContract } from './types';",
        "",
        "export const BASECG_VERSION = '2025-01-base-cg-retraite';",
        f"export const BASECG_EXTRACTED_COUNT = {len(contracts)};",
        "",
        "export const BASECG_CATALOG = [",
    ]
    for contract in contracts:
        lines.append(f"  {ts_json(contract)},")
    lines.extend([
        "] satisfies BaseCgRetraiteContract[];",
        "",
    ])
    BASECG_OUT.write_text("\n".join(lines), encoding="utf-8")


def download(url: str, path: Path) -> None:
    if path.exists():
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    urllib.request.urlretrieve(url, path)


def dataframe_to_table(code: str, rda_name: str, object_name: str) -> dict[str, Any]:
    rda_path = TMP_DIR / rda_name
    download(f"{CASDATASETS_RAW}/{rda_name}", rda_path)
    result = pyreadr.read_r(str(rda_path))
    frame = result[object_name]

    ages = [int(age) for age in frame["x"].tolist()]
    base = 100000
    if code == "TPRV93":
        return {
            "code": code,
            "type": "single",
            "base": base,
            "minAge": min(ages),
            "maxAge": max(ages),
            "lx": [int(round(value)) for value in frame["lx"].tolist()],
        }

    generations: dict[str, list[int]] = {}
    for column in frame.columns:
        if not str(column).startswith("lx"):
            continue
        year = str(column)[2:]
        values = [int(round(value)) for value in frame[column].tolist()]
        generations[year] = values

    return {
        "code": code,
        "type": "generation",
        "base": base,
        "minAge": min(ages),
        "maxAge": max(ages),
        "generations": generations,
    }


def write_mortality_table(filename: str, const_name: str, table: dict[str, Any]) -> None:
    MORTALITY_DIR.mkdir(parents=True, exist_ok=True)
    body = ts_json(table)
    path = MORTALITY_DIR / filename
    path.write_text(
        "\n".join(
            [
                "/* Fichier genere par tools/scripts/generate-per-transfert-data.py. */",
                "/* Source : CASdatasets, donnees freMortTables. */",
                "import type { MortalityTable } from './types';",
                "",
                f"export const {const_name} = {body} satisfies MortalityTable;",
                "",
            ]
        ),
        encoding="utf-8",
    )


def write_mortality() -> None:
    tables = {
        "TPRV93": dataframe_to_table("TPRV93", *MORTALITY_FILES["tprv93"]),
        "TGF05": dataframe_to_table("TGF05", *MORTALITY_FILES["tgf05"]),
        "TGH05": dataframe_to_table("TGH05", *MORTALITY_FILES["tgh05"]),
        "TPG93": dataframe_to_table("TPG93", *MORTALITY_FILES["tpg93"]),
    }
    write_mortality_table("tprv93.generated.ts", "TPRV93", tables["TPRV93"])
    write_mortality_table("tgf05.generated.ts", "TGF05", tables["TGF05"])
    write_mortality_table("tgh05.generated.ts", "TGH05", tables["TGH05"])
    write_mortality_table("tpg93.generated.ts", "TPG93", tables["TPG93"])


def main() -> int:
    contracts = extract_basecg()
    write_basecg(contracts)
    write_mortality()
    print(f"BASECG : {len(contracts)} contrats")
    print("Tables mortalite : TPRV93, TGF05, TGH05, TPG93")
    return 0


if __name__ == "__main__":
    sys.exit(main())
