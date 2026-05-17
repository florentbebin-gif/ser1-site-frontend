from __future__ import annotations

import json
import re
import sys
from pathlib import Path
from typing import Any

import openpyxl

ROOT = Path(__file__).resolve().parents[2]
CATALOG = ROOT / "src" / "data" / "basecg" / "catalog.generated.ts"

BASECG_ROWS = {
    "sourceId": 1,
    "compagnie": 4,
    "nomContrat": 5,
    "dateCommercialisation": 8,
    "nombreFonds": 11,
    "repartitionUcEuro": 12,
    "rendementFondsEuro": 13,
    "fraisVersements": 14,
    "fraisGestion": 15,
    "fraisArbitrage": 16,
    "fraisTransfertSortant": 17,
    "clauseBeneficiaire": 18,
    "garantiesComplementaires": 19,
    "ageLimiteLiquidation": 20,
    "sortieCapitalRetraite": 21,
    "fractionnementCapital": 22,
    "rachatLibre": 23,
    "tableConversionRente": 24,
    "tableGarantieAdhesion": 25,
    "tauxTechnique": 26,
    "fraisArrerages": 27,
    "annuitesGaranties": 28,
    "reversionPossible": 29,
    "reversionIncluse": 30,
    "renteEstimee": 31,
}


def normalize_cell(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        cleaned = re.sub(r"[ \t]+", " ", value.replace("\r\n", "\n").replace("\r", "\n")).strip()
        return cleaned or None
    if hasattr(value, "isoformat"):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def normalize_for_compare(value: Any) -> Any:
    if isinstance(value, str):
        return re.sub(r"[ \t]+", " ", value.replace("\r\n", "\n").replace("\r", "\n")).strip()
    return value


def slugify(value: str) -> str:
    lowered = value.lower()
    replacements = {
        "à": "a",
        "â": "a",
        "ä": "a",
        "ç": "c",
        "é": "e",
        "è": "e",
        "ê": "e",
        "ë": "e",
        "î": "i",
        "ï": "i",
        "ô": "o",
        "ö": "o",
        "ù": "u",
        "û": "u",
        "ü": "u",
        "ÿ": "y",
        "œ": "oe",
        "æ": "ae",
    }
    for src, dst in replacements.items():
        lowered = lowered.replace(src, dst)
    return re.sub(r"[^a-z0-9]+", "-", lowered).strip("-") or "contrat"


def detect_contract_type(company: str, name: str, table_label: str | None) -> str:
    haystack = f"{company} {name} {table_label or ''}".upper()
    if "PREFON" in haystack or "PRÉFON" in haystack:
        return "PER_POINTS"
    if "COREM" in haystack and "SYSTEME PAR POINTS" in haystack:
        return "PER_POINTS"
    if "ART83" in haystack or "ARTICLE 83" in haystack:
        return "ARTICLE83"
    if "MADELIN" in haystack:
        return "MADELIN"
    if "PERP" in haystack:
        return "PERP"
    if "PERCO" in haystack:
        return "PERCO"
    if "PERIN" in haystack or "PER INDIVIDUEL" in haystack or re.search(r"\bPER\b", haystack):
        return "PERIN"
    return "AUTRE"


def detect_compartment(contract_type: str, name: str) -> str:
    haystack = name.upper()
    if "NON DEDUIT" in haystack or "NON DÉDUIT" in haystack:
        return "C1_BIS"
    if contract_type == "ARTICLE83":
        return "C3"
    if contract_type == "PERCO":
        return "C2"
    return "C1"


def extract_rate(value: Any) -> float | None:
    if isinstance(value, (int, float)):
        number = float(value)
        return number if number <= 1 else number / 100
    if not isinstance(value, str):
        return None
    match = re.search(r"(\d+(?:[,.]\d+)?)\s*%", value)
    if not match:
        return None
    return float(match.group(1).replace(",", ".")) / 100


def resolve_xlsm_path() -> Path:
    candidates = sorted(ROOT.glob("*.xlsm"))
    if len(candidates) != 1:
        raise SystemExit("Définir BASECG_XLSM_PATH ou conserver un seul classeur .xlsm à la racine.")
    return candidates[0]


def extract_excel() -> list[dict[str, Any]]:
    workbook = openpyxl.load_workbook(resolve_xlsm_path(), data_only=True, read_only=False, keep_links=False)
    sheet = workbook["BASECG"]
    contracts: list[dict[str, Any]] = []
    for col in range(3, sheet.max_column + 1):
        raw = {key: normalize_cell(sheet.cell(row=row, column=col).value) for key, row in BASECG_ROWS.items()}
        if not raw["compagnie"] or not raw["nomContrat"]:
            continue
        company = str(raw["compagnie"])
        name = str(raw["nomContrat"])
        contract_type = detect_contract_type(company, name, raw.get("tableConversionRente"))
        contracts.append({
            "id": f"{slugify(company)}-{slugify(name)}-{col - 2}",
            "sourceId": str(raw["sourceId"] or f"Contrat N°{col - 2}"),
            "compagnie": company,
            "nomContrat": name,
            "typeContrat": contract_type,
            "perCompartment": detect_compartment(contract_type, name),
            "phaseEpargne": {
                "dateCommercialisation": raw["dateCommercialisation"],
                "nombreFonds": raw["nombreFonds"],
                "repartitionUcEuro": raw["repartitionUcEuro"],
                "rendementFondsEuro": raw["rendementFondsEuro"],
                "fraisVersements": raw["fraisVersements"],
                "fraisGestion": raw["fraisGestion"],
                "fraisArbitrage": raw["fraisArbitrage"],
                "fraisTransfertSortant": raw["fraisTransfertSortant"],
                "fraisTransfertSortantRate": extract_rate(raw["fraisTransfertSortant"]),
                "clauseBeneficiaire": raw["clauseBeneficiaire"],
                "garantiesComplementaires": raw["garantiesComplementaires"],
            },
            "phaseLiquidation": {
                "ageLimiteLiquidation": raw["ageLimiteLiquidation"],
                "sortieCapitalRetraite": raw["sortieCapitalRetraite"],
                "fractionnementCapital": raw["fractionnementCapital"],
                "rachatLibre": raw["rachatLibre"],
                "tableConversionRente": raw["tableConversionRente"],
                "tableGarantieAdhesion": raw["tableGarantieAdhesion"],
                "tauxTechnique": raw["tauxTechnique"],
                "fraisArrerages": raw["fraisArrerages"],
                "fraisArreragesRate": extract_rate(raw["fraisArrerages"]),
                "annuitesGaranties": raw["annuitesGaranties"],
                "reversionPossible": raw["reversionPossible"],
                "reversionIncluse": raw["reversionIncluse"],
                "renteEstimee": raw["renteEstimee"],
            },
        })
    return contracts


def extract_generated() -> list[dict[str, Any]]:
    text = CATALOG.read_text(encoding="utf-8")
    body = text.split("export const BASECG_CATALOG = [", 1)[1].split("] satisfies", 1)[0]
    contracts = []
    for line in body.splitlines():
        line = line.strip().rstrip(",")
        if line.startswith("{"):
            contracts.append(json.loads(line))
    return contracts


def main() -> int:
    expected = extract_excel()
    actual = extract_generated()
    actual_has_compartment = any("perCompartment" in contract for contract in actual)
    if not actual_has_compartment:
        for contract in expected:
            contract.pop("perCompartment", None)
    errors: list[str] = []
    if len(expected) != len(actual):
        errors.append(f"Nombre de contrats différent : Excel={len(expected)} / généré={len(actual)}")
    for index, (left, right) in enumerate(zip(expected, actual), start=1):
        if left["id"] != right["id"]:
            errors.append(f"Contrat {index}: id Excel={left['id']} / généré={right['id']}")
            continue
        left_json = json.dumps(left, ensure_ascii=False, sort_keys=True, default=normalize_for_compare)
        right_json = json.dumps(right, ensure_ascii=False, sort_keys=True, default=normalize_for_compare)
        if normalize_for_compare(left_json) != normalize_for_compare(right_json):
            errors.append(f"Contrat {left['id']}: divergence hors normalisation d’espaces")
    if errors:
        print("\n".join(errors[:30]), file=sys.stderr)
        if len(errors) > 30:
            print(f"... {len(errors) - 30} divergences supplémentaires", file=sys.stderr)
        return 1
    print(f"Audit BASECG OK : {len(actual)} contrats alignés avec l’onglet Excel BASECG.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
