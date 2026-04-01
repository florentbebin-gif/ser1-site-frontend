from __future__ import annotations

import json
import os
from copy import deepcopy
from datetime import date
from pathlib import Path
from typing import Any, Iterable
from urllib.error import HTTPError, URLError
from urllib.parse import urlencode
from urllib.request import Request, urlopen

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
OUTPUT_PATH = ROOT / "docs" / "settings-fiscaux-2026.xlsx"
SNAPSHOT_PATH = ROOT / ".tmp" / "settings-live-snapshot.json"
ENV_PATH = ROOT / ".env.local"
TODAY = date.today().isoformat()

COPY_COLUMNS = [
    "route",
    "section",
    "ui_label",
    "slot_ui",
    "annee_reference",
    "shared_group",
    "supabase_table",
    "json_path",
    "valeur_live_actuelle",
    "fallback_code",
    "valeur_cible",
    "unite",
    "source_officielle",
    "statut",
    "commentaire",
]

INVENTORY_COLUMNS = [
    "route",
    "section",
    "type",
    "shared_group",
    "supabase_table",
    "json_path",
    "editable",
    "consumer_status",
    "consumer_proof",
    "notes",
]

SOURCE_COLUMNS = [
    "famille",
    "champ_couvert",
    "source_url",
    "autorite",
    "date_verification",
    "notes",
]

USAGE_COLUMNS = [
    "ui_route",
    "shared_group",
    "source_table",
    "normalisation",
    "consumer",
    "preuve",
]

HEADER_FILL = PatternFill("solid", fgColor="17365D")
HEADER_FONT = Font(color="FFFFFF", bold=True)
STATUS_FILLS = {
    "a_mettre_a_jour": PatternFill("solid", fgColor="FFF2CC"),
    "stable_2026": PatternFill("solid", fgColor="E2F0D9"),
    "miroir_ui": PatternFill("solid", fgColor="D9EAF7"),
    "a_valider_metier": PatternFill("solid", fgColor="FCE4D6"),
}


def read_env_file(path: Path) -> dict[str, str]:
    values: dict[str, str] = {}
    if not path.exists():
        return values
    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        values[key.strip()] = value.strip().strip('"').strip("'")
    return values


def coalesce(*values: Any) -> Any:
    for value in values:
        if value not in (None, ""):
            return value
    return None


def first_dict(value: Any) -> dict[str, Any]:
    if isinstance(value, dict):
        return value
    if isinstance(value, list) and value and isinstance(value[0], dict):
        return value[0]
    return {}


def normalize_snapshot(raw: dict[str, Any]) -> dict[str, Any]:
    if "tax_settings" in raw:
        return raw

    tax_raw = raw.get("tax", {})
    ps_raw = raw.get("ps", {})
    fiscality_raw = raw.get("fiscality", {})
    pass_raw = raw.get("pass", {})

    tax_data = first_dict(tax_raw.get("data"))
    ps_data = first_dict(ps_raw.get("data"))
    fiscality_data = first_dict(fiscality_raw.get("data"))

    return {
        "tax_settings": {
            "updated_at": tax_data.get("updated_at"),
            "data": tax_data.get("data") or {},
        },
        "ps_settings": {
            "updated_at": ps_data.get("updated_at"),
            "data": ps_data.get("data") or {},
        },
        "fiscality_settings": {
            "updated_at": fiscality_data.get("updated_at"),
            "data": fiscality_data.get("data") or {},
        },
        "pass_history": {
            "data": pass_raw.get("data") or [],
        },
    }


def rest_get_json(base_url: str, service_key: str, table: str, params: dict[str, str]) -> Any:
    query = urlencode(params, safe=",()")
    url = f"{base_url.rstrip('/')}/rest/v1/{table}?{query}"
    request = Request(
        url,
        headers={
            "apikey": service_key,
            "Authorization": f"Bearer {service_key}",
            "Accept": "application/json",
        },
    )
    with urlopen(request, timeout=30) as response:
        return json.loads(response.read().decode("utf-8"))


def fetch_live_snapshot() -> dict[str, Any]:
    env_values = read_env_file(ENV_PATH)
    supabase_url = coalesce(
        os.environ.get("SUPABASE_URL"),
        os.environ.get("VITE_SUPABASE_URL"),
        env_values.get("SUPABASE_URL"),
        env_values.get("VITE_SUPABASE_URL"),
    )
    service_key = coalesce(
        os.environ.get("SUPABASE_SERVICE_ROLE_KEY"),
        os.environ.get("SUPABASE_ANON_KEY"),
        os.environ.get("VITE_SUPABASE_ANON_KEY"),
        env_values.get("SUPABASE_SERVICE_ROLE_KEY"),
        env_values.get("SUPABASE_ANON_KEY"),
        env_values.get("VITE_SUPABASE_ANON_KEY"),
    )
    if not supabase_url or not service_key:
        raise RuntimeError("Supabase credentials are missing.")

    tax_rows = rest_get_json(
        supabase_url,
        service_key,
        "tax_settings",
        {"select": "id,data,updated_at", "id": "eq.1"},
    )
    ps_rows = rest_get_json(
        supabase_url,
        service_key,
        "ps_settings",
        {"select": "id,data,updated_at", "id": "eq.1"},
    )
    fiscality_rows = rest_get_json(
        supabase_url,
        service_key,
        "fiscality_settings",
        {"select": "id,data,updated_at", "id": "eq.1"},
    )
    pass_rows = rest_get_json(
        supabase_url,
        service_key,
        "pass_history",
        {"select": "year,pass_amount,updated_at", "order": "year.asc"},
    )

    snapshot = {
        "tax_settings": first_dict(tax_rows),
        "ps_settings": first_dict(ps_rows),
        "fiscality_settings": first_dict(fiscality_rows),
        "pass_history": {"data": pass_rows or []},
    }
    SNAPSHOT_PATH.parent.mkdir(parents=True, exist_ok=True)
    SNAPSHOT_PATH.write_text(
        json.dumps(snapshot, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return snapshot


def load_snapshot() -> dict[str, Any]:
    try:
        return fetch_live_snapshot()
    except (HTTPError, URLError, RuntimeError, TimeoutError):
        if SNAPSHOT_PATH.exists():
            raw = json.loads(SNAPSHOT_PATH.read_text(encoding="utf-8"))
            return normalize_snapshot(raw)
        raise


def get_path(data: Any, *path: Any) -> Any:
    current = data
    for part in path:
        if current is None:
            return None
        if isinstance(part, int):
            if not isinstance(current, list) or part >= len(current):
                return None
            current = current[part]
            continue
        if not isinstance(current, dict):
            return None
        current = current.get(part)
    return current


def stringify_path(parts: Iterable[Any]) -> str:
    text = ""
    for index, part in enumerate(parts):
        if isinstance(part, int):
            text += f"[{part}]"
        else:
            if index > 0:
                text += "."
            text += str(part)
    return text


def normalize_value(value: Any) -> Any:
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def values_equal(left: Any, right: Any) -> bool:
    return normalize_value(left) == normalize_value(right)


def deep_merge(base: Any, override: Any) -> Any:
    if isinstance(base, dict) and isinstance(override, dict):
        merged = {key: deepcopy(value) for key, value in base.items()}
        for key, value in override.items():
            merged[key] = deep_merge(merged.get(key), value)
        return merged
    if isinstance(base, list) and isinstance(override, list):
        return deepcopy(override)
    return deepcopy(override if override is not None else base)


def build_live_map(rows: list[dict[str, Any]]) -> dict[int, Any]:
    live_map: dict[int, Any] = {}
    for row in rows:
        year = row.get("year")
        if year is None:
            continue
        live_map[int(year)] = row.get("pass_amount")
    return live_map


def with_status_fill(cell: Any, status: str) -> None:
    fill = STATUS_FILLS.get(status)
    if fill:
        cell.fill = fill


FALLBACK_TAX_SETTINGS = {
    "incomeTax": {
        "currentYearLabel": "2025 (revenus 2024)",
        "previousYearLabel": "2024 (revenus 2023)",
        "scaleCurrent": [
            {"from": 0, "to": 11497, "rate": 0},
            {"from": 11498, "to": 29315, "rate": 11},
            {"from": 29316, "to": 83823, "rate": 30},
            {"from": 83824, "to": 180294, "rate": 41},
            {"from": 180295, "to": None, "rate": 45},
        ],
        "scalePrevious": [
            {"from": 0, "to": 11294, "rate": 0},
            {"from": 11295, "to": 28797, "rate": 11},
            {"from": 28798, "to": 82341, "rate": 30},
            {"from": 82342, "to": 177106, "rate": 41},
            {"from": 177107, "to": None, "rate": 45},
        ],
        "quotientFamily": {
            "current": {
                "plafondPartSup": 1791,
                "plafondParentIsoleDeuxPremieresParts": 4224,
            },
            "previous": {
                "plafondPartSup": 1791,
                "plafondParentIsoleDeuxPremieresParts": 4224,
            },
        },
        "decote": {
            "current": {
                "triggerSingle": 1964,
                "triggerCouple": 3248,
                "amountSingle": 889,
                "amountCouple": 1470,
                "ratePercent": 45.25,
            },
            "previous": {
                "triggerSingle": 1964,
                "triggerCouple": 3248,
                "amountSingle": 889,
                "amountCouple": 1470,
                "ratePercent": 45.25,
            },
        },
        "abat10": {
            "current": {"plafond": 14426, "plancher": 504},
            "previous": {"plafond": 14171, "plancher": 495},
            "retireesCurrent": {"plafond": 4399, "plancher": 450},
            "retireesPrevious": {"plafond": 4321, "plancher": 442},
        },
        "domAbatement": {
            "current": {
                "gmr": {"ratePercent": 30, "cap": 2450},
                "guyane": {"ratePercent": 40, "cap": 4050},
            },
            "previous": {
                "gmr": {"ratePercent": 30, "cap": 2450},
                "guyane": {"ratePercent": 40, "cap": 4050},
            },
        },
    },
    "pfu": {
        "current": {"rateIR": 12.8, "rateSocial": 17.2, "rateTotal": 30.0},
        "previous": {"rateIR": 12.8, "rateSocial": 17.2, "rateTotal": 30.0},
    },
    "cehr": {
        "current": {
            "single": [
                {"from": 250000, "to": 500000, "rate": 3},
                {"from": 500000, "to": None, "rate": 4},
            ],
            "couple": [
                {"from": 500000, "to": 1000000, "rate": 3},
                {"from": 1000000, "to": None, "rate": 4},
            ],
        },
        "previous": {
            "single": [
                {"from": 250000, "to": 500000, "rate": 3},
                {"from": 500000, "to": None, "rate": 4},
            ],
            "couple": [
                {"from": 500000, "to": 1000000, "rate": 3},
                {"from": 1000000, "to": None, "rate": 4},
            ],
        },
    },
    "cdhr": {
        "current": {"minEffectiveRate": 20, "thresholdSingle": 250000, "thresholdCouple": 500000},
        "previous": {"minEffectiveRate": 20, "thresholdSingle": 250000, "thresholdCouple": 500000},
    },
    "corporateTax": {
        "current": {"normalRate": 25, "reducedRate": 15, "reducedThreshold": 42500},
        "previous": {"normalRate": 25, "reducedRate": 15, "reducedThreshold": 42500},
    },
    "dmtg": {
        "ligneDirecte": {
            "abattement": 100000,
            "scale": [
                {"from": 0, "to": 8072, "rate": 5},
                {"from": 8072, "to": 12109, "rate": 10},
                {"from": 12109, "to": 15932, "rate": 15},
                {"from": 15932, "to": 552324, "rate": 20},
                {"from": 552324, "to": 902838, "rate": 30},
                {"from": 902838, "to": 1805677, "rate": 40},
                {"from": 1805677, "to": None, "rate": 45},
            ],
        },
        "frereSoeur": {
            "abattement": 15932,
            "scale": [
                {"from": 0, "to": 24430, "rate": 35},
                {"from": 24430, "to": None, "rate": 45},
            ],
        },
        "neveuNiece": {
            "abattement": 7967,
            "scale": [{"from": 0, "to": None, "rate": 55}],
        },
        "autre": {
            "abattement": 1594,
            "scale": [{"from": 0, "to": None, "rate": 60}],
        },
    },
    "donation": {
        "rappelFiscalAnnees": 15,
        "donFamilial790G": {
            "montant": 31865,
            "conditions": "Donateur < 80 ans, donataire majeur",
        },
        "donManuel": {"abattementRenouvellement": 15},
    },
}

FALLBACK_PS_SETTINGS = {
    "labels": {
        "currentYearLabel": "2025 (RFR 2023 & Avis IR 2024)",
        "previousYearLabel": "2024 (RFR 2022 & Avis IR 2023)",
    },
    "patrimony": {
        "current": {"totalRate": 17.2, "csgDeductibleRate": 6.8},
        "previous": {"totalRate": 17.2, "csgDeductibleRate": 6.8},
    },
    "retirement": {
        "current": {
            "brackets": [
                {"label": "Exoneration", "rfrMin1Part": 0, "rfrMax1Part": 11432, "csgRate": 0, "crdsRate": 0, "casaRate": 0, "maladieRate": 0, "totalRate": 0, "csgDeductibleRate": 0},
                {"label": "Taux reduit", "rfrMin1Part": 11433, "rfrMax1Part": 14944, "csgRate": 3.8, "crdsRate": 0.5, "casaRate": 0, "maladieRate": 0, "totalRate": 4.3, "csgDeductibleRate": 3.8},
                {"label": "Taux median", "rfrMin1Part": 14945, "rfrMax1Part": 23193, "csgRate": 6.6, "crdsRate": 0.5, "casaRate": 0.3, "maladieRate": 1.0, "totalRate": 8.4, "csgDeductibleRate": 4.2},
                {"label": "Taux normal", "rfrMin1Part": 23193, "rfrMax1Part": None, "csgRate": 8.3, "crdsRate": 0.5, "casaRate": 0.3, "maladieRate": 1.0, "totalRate": 10.1, "csgDeductibleRate": 5.9},
            ],
        },
        "previous": {
            "brackets": [
                {"label": "Exoneration", "rfrMin1Part": 0, "rfrMax1Part": 11432, "csgRate": 0, "crdsRate": 0, "casaRate": 0, "maladieRate": 0, "totalRate": 0, "csgDeductibleRate": 0},
                {"label": "Taux reduit", "rfrMin1Part": 11433, "rfrMax1Part": 14944, "csgRate": 3.8, "crdsRate": 0.5, "casaRate": 0, "maladieRate": 0, "totalRate": 4.3, "csgDeductibleRate": 3.8},
                {"label": "Taux median", "rfrMin1Part": 14945, "rfrMax1Part": 23193, "csgRate": 6.6, "crdsRate": 0.5, "casaRate": 0.3, "maladieRate": 1.0, "totalRate": 8.4, "csgDeductibleRate": 4.2},
                {"label": "Taux normal", "rfrMin1Part": 23193, "rfrMax1Part": None, "csgRate": 8.3, "crdsRate": 0.5, "casaRate": 0.3, "maladieRate": 1.0, "totalRate": 10.1, "csgDeductibleRate": 5.9},
            ],
        },
    },
    "retirementThresholds": {
        "current": {
            "metropole": {"rfrMaxExemption1Part": 12817, "rfrMaxReduced1Part": 16755, "rfrMaxMedian1Part": 26004, "incrementQuarterExemption": 1711, "incrementQuarterReduced": 2237, "incrementQuarterMedian": 3471},
            "gmr": {"rfrMaxExemption1Part": 15164, "rfrMaxReduced1Part": 18331, "rfrMaxMedian1Part": 26004, "incrementQuarterExemption": 1882, "incrementQuarterReduced": 2459, "incrementQuarterMedian": 3471},
            "guyane": {"rfrMaxExemption1Part": 15856, "rfrMaxReduced1Part": 19200, "rfrMaxMedian1Part": 26004, "incrementQuarterExemption": 1968, "incrementQuarterReduced": 2572, "incrementQuarterMedian": 3471},
        },
        "previous": {
            "metropole": {"rfrMaxExemption1Part": 12230, "rfrMaxReduced1Part": 15988, "rfrMaxMedian1Part": 24812, "incrementQuarterExemption": 1633, "incrementQuarterReduced": 2135, "incrementQuarterMedian": 3312},
            "gmr": {"rfrMaxExemption1Part": 14469, "rfrMaxReduced1Part": 17491, "rfrMaxMedian1Part": 24812, "incrementQuarterExemption": 1633, "incrementQuarterReduced": 2135, "incrementQuarterMedian": 3312},
            "guyane": {"rfrMaxExemption1Part": 15130, "rfrMaxReduced1Part": 18321, "rfrMaxMedian1Part": 24812, "incrementQuarterExemption": 1633, "incrementQuarterReduced": 2135, "incrementQuarterMedian": 3312},
        },
    },
}

FALLBACK_FISCALITY_SETTINGS = {
    "assuranceVie": {
        "deces": {
            "agePivotPrimes": 70,
            "primesApres1998": {
                "allowancePerBeneficiary": 152500,
                "brackets": [
                    {"upTo": 700000, "ratePercent": 20},
                    {"upTo": None, "ratePercent": 31.25},
                ],
            },
            "apres70ans": {
                "globalAllowance": 30500,
                "taxationMode": "dmtg",
            },
        }
    }
}

FALLBACK_PASS_HISTORY = {
    2019: 40524,
    2020: 41136,
    2021: 41136,
    2022: 41136,
    2023: 43992,
    2024: 46368,
    2025: 47100,
}

SOURCE_URLS = {
    "baseline_live": "Supabase live snapshot 2026-04-01",
    "millesime": "Convention SER1 - decalage du slot current vers previous",
    "ir_2026": "https://www.service-public.gouv.fr/particuliers/actualites/A18045",
    "ir_reductions": "https://www.service-public.gouv.fr/particuliers/vosdroits/F34328",
    "pfu_2026": "https://entreprendre.service-public.gouv.fr/actualites/A18796",
    "ps_retraite_2026": "https://www.service-public.gouv.fr/particuliers/vosdroits/F2971",
    "ps_dom_2026": "https://legislation.lassuranceretraite.fr/Pdf/instruction_ministerielle_11122025.pdf",
    "pass_2026": "https://www.service-public.gouv.fr/particuliers/actualites/A15386",
    "dmtg": "https://www.service-public.gouv.fr/particuliers/vosdroits/F14198",
    "donation": "https://www.service-public.gouv.fr/particuliers/vosdroits/F10203",
    "donation_790g": "https://www.service-public.gouv.fr/particuliers/vosdroits/F36656",
    "av_deces": "https://bofip.impots.gouv.fr/bofip/1335-PGP.html/identifiant=BOI-TCAS-AUT-60-20120912",
}

OFFICIAL_2026_OVERRIDES = {
    "tax": {
        "incomeTax": {
            "currentYearLabel": "2026 (revenus 2025)",
            "previousYearLabel": "2025 (revenus 2024)",
            "scaleCurrent": [
                {"from": 0, "to": 11600, "rate": 0},
                {"from": 11601, "to": 29579, "rate": 11},
                {"from": 29580, "to": 84577, "rate": 30},
                {"from": 84578, "to": 181917, "rate": 41},
                {"from": 181918, "to": None, "rate": 45},
            ],
            "quotientFamily": {
                "current": {
                    "plafondPartSup": 1807,
                    "plafondParentIsoleDeuxPremieresParts": 4262,
                }
            },
            "decote": {
                "current": {
                    "triggerSingle": 1982,
                    "triggerCouple": 3277,
                    "amountSingle": 897,
                    "amountCouple": 1483,
                    "ratePercent": 45.25,
                }
            },
            "abat10": {
                "current": {"plafond": 14555, "plancher": 509},
                "retireesCurrent": {"plafond": 4439, "plancher": 454},
            },
        },
        "pfu": {
            "current": {"rateIR": 12.8, "rateSocial": 18.6, "rateTotal": 31.4},
        },
    },
    "ps": {
        "labels": {
            "currentYearLabel": "2026 (RFR 2024 & Avis IR 2025)",
            "previousYearLabel": "2025 (RFR 2023 & Avis IR 2024)",
        },
        "patrimony": {
            "current": {"totalRate": 18.6, "csgDeductibleRate": 6.8},
        },
        "retirement": {
            "current": {
                "brackets": [
                    {"label": "Exoneration", "rfrMin1Part": 0, "rfrMax1Part": 13048, "csgRate": 0, "crdsRate": 0, "casaRate": 0, "maladieRate": 0, "totalRate": 0, "csgDeductibleRate": 0},
                    {"label": "Taux reduit", "rfrMin1Part": 13049, "rfrMax1Part": 17057, "csgRate": 3.8, "crdsRate": 0.5, "casaRate": 0, "maladieRate": 0, "totalRate": 4.3, "csgDeductibleRate": 3.8},
                    {"label": "Taux median", "rfrMin1Part": 17058, "rfrMax1Part": 26472, "csgRate": 6.6, "crdsRate": 0.5, "casaRate": 0.3, "maladieRate": 1.0, "totalRate": 8.4, "csgDeductibleRate": 4.2},
                    {"label": "Taux normal", "rfrMin1Part": 26473, "rfrMax1Part": None, "csgRate": 8.3, "crdsRate": 0.5, "casaRate": 0.3, "maladieRate": 1.0, "totalRate": 10.1, "csgDeductibleRate": 5.9},
                ],
            },
        },
        "retirementThresholds": {
            "current": {
                "metropole": {"rfrMaxExemption1Part": 13048, "rfrMaxReduced1Part": 17057, "rfrMaxMedian1Part": 26472, "incrementQuarterExemption": 1742, "incrementQuarterReduced": 2278, "incrementQuarterMedian": 3533},
                "gmr": {"rfrMaxExemption1Part": 15437, "rfrMaxReduced1Part": 18661, "rfrMaxMedian1Part": 26472, "incrementQuarterExemption": 1916, "incrementQuarterReduced": 2504, "incrementQuarterMedian": 3533},
                "guyane": {"rfrMaxExemption1Part": 16141, "rfrMaxReduced1Part": 19546, "rfrMaxMedian1Part": 26472, "incrementQuarterExemption": 2003, "incrementQuarterReduced": 2619, "incrementQuarterMedian": 3533},
            }
        },
    },
    "fiscality": {
        "assuranceVie": {
            "deces": {
                "agePivotPrimes": 70,
                "primesApres1998": {
                    "allowancePerBeneficiary": 152500,
                    "brackets": [
                        {"upTo": 852500, "ratePercent": 20},
                        {"upTo": None, "ratePercent": 31.25},
                    ],
                },
                "apres70ans": {"globalAllowance": 30500, "taxationMode": "dmtg"},
            }
        }
    },
    "pass_history": {
        2026: 48060,
    },
}

DMTG_CATEGORIES = [
    ("ligneDirecte", "Ligne directe (enfants, petits-enfants)", "Abattement par enfant"),
    ("frereSoeur", "Freres et soeurs", "Abattement frere/soeur"),
    ("neveuNiece", "Neveux et nieces", "Abattement neveu/niece"),
    ("autre", "Autres (non-parents)", "Abattement par defaut"),
]

DOM_ZONES = [
    ("gmr", "Guadeloupe / Martinique / Reunion"),
    ("guyane", "Guyane / Mayotte"),
]

RETIREMENT_LABELS = [
    "Exoneration",
    "Taux reduit",
    "Taux median",
    "Taux normal",
]

RETIREMENT_FIELDS = [
    ("rfrMin1Part", "RFR min (1 part)", "EUR"),
    ("rfrMax1Part", "RFR max (1 part)", "EUR"),
    ("csgRate", "CSG %", "%"),
    ("crdsRate", "CRDS %", "%"),
    ("casaRate", "CASA %", "%"),
    ("maladieRate", "Maladie %", "%"),
    ("totalRate", "Total %", "%"),
    ("csgDeductibleRate", "CSG ded. %", "%"),
]

THRESHOLD_REGIONS = [
    ("metropole", "Residence en metropole"),
    ("gmr", "Residence en Martinique, Guadeloupe, Reunion, Saint-Barthelemy, Saint-Martin"),
    ("guyane", "Residence en Guyane"),
]

THRESHOLD_FIELDS = [
    ("rfrMaxExemption1Part", "Plafond exoneration (1 part)"),
    ("rfrMaxReduced1Part", "Plafond taux reduit (1 part)"),
    ("rfrMaxMedian1Part", "Plafond taux median (1 part)"),
    ("incrementQuarterExemption", "Majoration par quart - exoneration"),
    ("incrementQuarterReduced", "Majoration par quart - taux reduit"),
    ("incrementQuarterMedian", "Majoration par quart - taux median"),
]

READ_ONLY_SECTIONS = [
    ("/settings/dmtg-succession", "Reserve hereditaire & droits du conjoint"),
    ("/settings/dmtg-succession", "Regimes matrimoniaux & PACS"),
    ("/settings/dmtg-succession", "Liberalites"),
    ("/settings/dmtg-succession", "Avantages matrimoniaux"),
]

DMTG_PROOF = (
    "SettingsImpots.tsx:117-165,327-333; "
    "SettingsDmtgSuccession.tsx:78-80,208-210,252-258; "
    "useFiscalContext.ts:120-152; "
    "successionFiscalContext.ts:101-115; "
    "usePlacementSimulatorController.ts:102-104"
)

GROUP_CONSUMERS = {
    "incomeTax": {
        "status": "utilise",
        "proof": (
            "useFiscalContext.ts:129-135 -> "
            "IrSimulatorContainer.tsx:286-331 -> "
            "engine/ir/compute.ts:82-177"
        ),
    },
    "domAbatement": {
        "status": "utilise",
        "proof": (
            "IrSimulatorContainer.tsx:286-331 -> "
            "engine/ir/compute.ts:161-177"
        ),
    },
    "pfu": {
        "status": "utilise",
        "proof": (
            "IrSimulatorContainer.tsx:317-331 -> "
            "engine/ir/compute.ts:176-177; "
            "fiscalitySettingsMigrator.ts:39-41"
        ),
    },
    "cehr_cdhr": {
        "status": "utilise",
        "proof": (
            "IrSimulatorContainer.tsx:286-331 -> "
            "engine/ir/compute.ts:97-102,197-204"
        ),
    },
    "corporateTax": {
        "status": "aucun match runtime trouve dans la recherche actuelle",
        "proof": (
            "rg \"corporateTax\" src -> settingsDefaults.ts, "
            "ImpotsISSection.tsx, dmtgValidators.ts"
        ),
    },
    "dmtg": {
        "status": "utilise",
        "proof": DMTG_PROOF,
    },
    "pass_history": {
        "status": "utilise",
        "proof": (
            "PassHistoryAccordion.tsx:34-39,92; "
            "useFiscalContext.ts:154-155; "
            "PerPotentielSimulator.tsx:110; "
            "usePerPotentiel.ts:187"
        ),
    },
    "patrimony": {
        "status": "utilise",
        "proof": (
            "useFiscalContext.ts:138-140; "
            "IrSimulatorContainer.tsx:326-331; "
            "engine/placement/fiscalParams.ts:57-60; "
            "PlacementTransmissionSection.tsx:211"
        ),
    },
    "retirement": {
        "status": "aucun match runtime trouve dans la recherche actuelle",
        "proof": (
            "rg \"retirement\\.\" src -> settingsDefaults.ts, "
            "SettingsPrelevements.tsx, validators"
        ),
    },
    "retirementThresholds": {
        "status": "aucun match runtime trouve dans la recherche actuelle",
        "proof": (
            "rg \"retirementThresholds\" src -> settingsDefaults.ts, "
            "SettingsPrelevements.tsx, SeuilsYearPeriod.tsx, validators"
        ),
    },
    "donation": {
        "status": "utilise",
        "proof": (
            "successionFiscalContext.ts:107-127; "
            "successionPatrimonial.ts:223-230"
        ),
    },
    "av_deces": {
        "status": "utilise",
        "proof": (
            "successionFiscalContext.ts:129-144; "
            "engine/placement/fiscalParams.ts:72-93"
        ),
    },
    "ps_labels": {
        "status": "visible dans la page settings",
        "proof": (
            "PrelevementsPatrimoineSection.tsx:66-92; "
            "PrelevementsRetraitesSection.tsx:58-100; "
            "PrelevementsSeuilsSection.tsx:69-86"
        ),
    },
}


def build_source_rows() -> list[dict[str, Any]]:
    return [
        {
            "famille": "baseline interne",
            "champ_couvert": "valeurs live tax_settings / ps_settings / fiscality_settings / pass_history",
            "source_url": SOURCE_URLS["baseline_live"],
            "autorite": "Supabase workspace SER1",
            "date_verification": TODAY,
            "notes": "Snapshot live prioritaire pour la colonne valeur_live_actuelle et pour le slot previous 2025.",
        },
        {
            "famille": "millesime",
            "champ_couvert": "libelles currentYearLabel / previousYearLabel",
            "source_url": SOURCE_URLS["millesime"],
            "autorite": "Convention SER1",
            "date_verification": TODAY,
            "notes": "Le slot previous 2025 reprend l'ancien slot current 2025; le slot current devient 2026.",
        },
        {
            "famille": "impot revenu",
            "champ_couvert": "bareme IR 2026",
            "source_url": SOURCE_URLS["ir_2026"],
            "autorite": "Service-Public.fr",
            "date_verification": TODAY,
            "notes": "Tranches 2026 pour les revenus 2025.",
        },
        {
            "famille": "impot revenu",
            "champ_couvert": "quotient familial, decote, abattements 10 %",
            "source_url": SOURCE_URLS["ir_reductions"],
            "autorite": "Service-Public.fr",
            "date_verification": TODAY,
            "notes": "Utilise pour le quotient familial, la decote et les abattements 10 % 2026.",
        },
        {
            "famille": "PFU",
            "champ_couvert": "PFU 2026",
            "source_url": SOURCE_URLS["pfu_2026"],
            "autorite": "Service-Public.fr",
            "date_verification": TODAY,
            "notes": "PFU 2026 a 31,4 % = 12,8 % IR + 18,6 % PS.",
        },
        {
            "famille": "prelevements sociaux",
            "champ_couvert": "retraites 2026 metropole",
            "source_url": SOURCE_URLS["ps_retraite_2026"],
            "autorite": "Service-Public.fr",
            "date_verification": TODAY,
            "notes": "F2971 pour les seuils metropole et les taux generiques sur pensions.",
        },
        {
            "famille": "prelevements sociaux",
            "champ_couvert": "seuils retraites 2026 GMR / Guyane",
            "source_url": SOURCE_URLS["ps_dom_2026"],
            "autorite": "Ministere / Assurance retraite",
            "date_verification": TODAY,
            "notes": "Instruction du 2025-12-11 pour les seuils DOM 2026.",
        },
        {
            "famille": "PASS",
            "champ_couvert": "PASS 2026",
            "source_url": SOURCE_URLS["pass_2026"],
            "autorite": "Service-Public.fr",
            "date_verification": TODAY,
            "notes": "PASS 2026 = 48 060 EUR.",
        },
        {
            "famille": "DMTG",
            "champ_couvert": "abattements et baremes succession",
            "source_url": SOURCE_URLS["dmtg"],
            "autorite": "Service-Public.fr",
            "date_verification": TODAY,
            "notes": "Valeurs stables 2026 pour les categories ligne directe, frere/soeur, neveu/niece et autre.",
        },
        {
            "famille": "donation",
            "champ_couvert": "rappel fiscal et donations",
            "source_url": SOURCE_URLS["donation"],
            "autorite": "Service-Public.fr",
            "date_verification": TODAY,
            "notes": "Rappel fiscal 15 ans. Utiliser aussi le focus 790 G.",
        },
        {
            "famille": "donation",
            "champ_couvert": "don familial de somme d'argent 790 G",
            "source_url": SOURCE_URLS["donation_790g"],
            "autorite": "Service-Public.fr",
            "date_verification": TODAY,
            "notes": "Abattement specifique 31 865 EUR.",
        },
        {
            "famille": "assurance-vie deces",
            "champ_couvert": "990 I / 757 B",
            "source_url": SOURCE_URLS["av_deces"],
            "autorite": "BOFiP",
            "date_verification": TODAY,
            "notes": "Le premier seuil est stocke cumule dans l'UI 852 500 = 152 500 + 700 000.",
        },
    ]


def build_dmtg_usage_rows() -> list[dict[str, Any]]:
    return [
        {
            "ui_route": "/settings/impots",
            "shared_group": "tax_settings.dmtg.*",
            "source_table": "tax_settings",
            "normalisation": "SettingsImpots charge et sauvegarde tax_settings.id=1",
            "consumer": "page admin impots",
            "preuve": "SettingsImpots.tsx:117-165,327-333",
        },
        {
            "ui_route": "/settings/dmtg-succession",
            "shared_group": "tax_settings.dmtg.*",
            "source_table": "tax_settings",
            "normalisation": "SettingsDmtgSuccession charge tax_settings.id=1 et sauvegarde avec fiscality_settings",
            "consumer": "page admin DMTG & Succession",
            "preuve": "SettingsDmtgSuccession.tsx:78-80,208-210,252-258",
        },
        {
            "ui_route": "runtime",
            "shared_group": "tax_settings.dmtg.*",
            "source_table": "tax_settings",
            "normalisation": "useFiscalContext normalise tax.dmtg en fiscalContext.dmtgSettings + dmtgScaleLigneDirecte",
            "consumer": "tous consommateurs via fiscalContext",
            "preuve": "useFiscalContext.ts:120-152",
        },
        {
            "ui_route": "/sim/succession",
            "shared_group": "tax_settings.dmtg.*",
            "source_table": "tax_settings",
            "normalisation": "buildSuccessionFiscalSnapshot lit fiscalContext.dmtgSettings",
            "consumer": "simulateur succession",
            "preuve": "successionFiscalContext.ts:101-115",
        },
        {
            "ui_route": "/sim/placement",
            "shared_group": "tax_settings.dmtg.ligneDirecte.scale",
            "source_table": "tax_settings",
            "normalisation": "usePlacementSimulatorController lit fiscalContext.dmtgScaleLigneDirecte",
            "consumer": "simulateur placement",
            "preuve": "usePlacementSimulatorController.ts:102-104",
        },
        {
            "ui_route": "/sim/placement",
            "shared_group": "tax_settings.dmtg.*",
            "source_table": "tax_settings",
            "normalisation": "usePlacementSettings extrait dmtgAbattementLigneDirecte et dmtgScale pour le moteur",
            "consumer": "moteur placement",
            "preuve": "usePlacementSettings.ts:138-145",
        },
    ]


def status_for_row(
    live_value: Any,
    target_value: Any,
    *,
    mirror: bool = False,
    metier_validation: bool = False,
) -> str:
    if mirror:
        return "miroir_ui"
    if metier_validation:
        return "a_valider_metier"
    if values_equal(live_value, target_value):
        return "stable_2026"
    return "a_mettre_a_jour"


def build_copy_row(
    *,
    route: str,
    section: str,
    ui_label: str,
    slot_ui: str,
    annee_reference: Any,
    shared_group: str,
    supabase_table: str,
    json_path: str,
    valeur_live_actuelle: Any,
    fallback_code: Any,
    valeur_cible: Any,
    unite: str,
    source_officielle: str,
    commentaire: str = "",
    mirror: bool = False,
    metier_validation: bool = False,
) -> dict[str, Any]:
    return {
        "route": route,
        "section": section,
        "ui_label": ui_label,
        "slot_ui": slot_ui,
        "annee_reference": annee_reference,
        "shared_group": shared_group,
        "supabase_table": supabase_table,
        "json_path": json_path,
        "valeur_live_actuelle": normalize_value(valeur_live_actuelle),
        "fallback_code": normalize_value(fallback_code),
        "valeur_cible": normalize_value(valeur_cible),
        "unite": unite,
        "source_officielle": source_officielle,
        "statut": status_for_row(
            valeur_live_actuelle,
            valeur_cible,
            mirror=mirror,
            metier_validation=metier_validation,
        ),
        "commentaire": commentaire,
    }


def build_annual_target(
    live_root: dict[str, Any],
    fallback_root: dict[str, Any],
    current_override: dict[str, Any],
    current_path: list[Any],
    previous_path: list[Any],
) -> tuple[Any, Any]:
    current_target = coalesce(
        get_path(current_override, *current_path),
        get_path(live_root, *current_path),
        get_path(fallback_root, *current_path),
    )
    previous_target = coalesce(
        get_path(live_root, *current_path),
        get_path(fallback_root, *current_path),
    )
    current_live = coalesce(
        get_path(live_root, *current_path),
        get_path(fallback_root, *current_path),
    )
    previous_live = coalesce(
        get_path(live_root, *previous_path),
        get_path(fallback_root, *previous_path),
    )
    return (current_live, current_target), (previous_live, previous_target)


def append_dmtg_rows(
    rows: list[dict[str, Any]],
    *,
    route: str,
    tax_live: dict[str, Any],
    tax_fallback: dict[str, Any],
    section: str,
) -> None:
    for category_key, category_title, abattement_label in DMTG_CATEGORIES:
        shared_prefix = f"tax_settings.dmtg.{category_key}"
        abattement_path = ["dmtg", category_key, "abattement"]
        rows.append(
            build_copy_row(
                route=route,
                section=section,
                ui_label=f"{category_title} - {abattement_label}",
                slot_ui="unique",
                annee_reference=2026,
                shared_group=f"{shared_prefix}.abattement",
                supabase_table="tax_settings",
                json_path=stringify_path(abattement_path),
                valeur_live_actuelle=coalesce(get_path(tax_live, *abattement_path), get_path(tax_fallback, *abattement_path)),
                fallback_code=get_path(tax_fallback, *abattement_path),
                valeur_cible=coalesce(get_path(tax_live, *abattement_path), get_path(tax_fallback, *abattement_path)),
                unite="EUR",
                source_officielle=SOURCE_URLS["dmtg"],
                commentaire="Champ miroir avec l'autre page DMTG; meme source tax_settings.dmtg.",
                mirror=True,
            )
        )
        scale_live = coalesce(get_path(tax_live, "dmtg", category_key, "scale"), get_path(tax_fallback, "dmtg", category_key, "scale"), [])
        scale_fallback = coalesce(get_path(tax_fallback, "dmtg", category_key, "scale"), [])
        for index, live_row in enumerate(scale_live):
            fallback_row = scale_fallback[index] if index < len(scale_fallback) else {}
            for field_key, field_label, unit in [
                ("from", "De (EUR)", "EUR"),
                ("to", "A (EUR)", "EUR"),
                ("rate", "Taux %", "%"),
            ]:
                target_value = live_row.get(field_key)
                comment = "Laisser vide pour borne ouverte." if field_key == "to" and target_value is None else "Champ miroir avec l'autre page DMTG; meme source tax_settings.dmtg."
                rows.append(
                    build_copy_row(
                        route=route,
                        section=section,
                        ui_label=f"{category_title} - Tranche {index + 1} - {field_label}",
                        slot_ui="unique",
                        annee_reference=2026,
                        shared_group=f"{shared_prefix}.scale[{index}].{field_key}",
                        supabase_table="tax_settings",
                        json_path=stringify_path(["dmtg", category_key, "scale", index, field_key]),
                        valeur_live_actuelle=live_row.get(field_key),
                        fallback_code=fallback_row.get(field_key),
                        valeur_cible=target_value,
                        unite=unit,
                        source_officielle=SOURCE_URLS["dmtg"],
                        commentaire=comment,
                        mirror=True,
                    )
                )


def build_impots_rows(snapshot: dict[str, Any]) -> list[dict[str, Any]]:
    tax_live = deep_merge(FALLBACK_TAX_SETTINGS, snapshot["tax_settings"].get("data") or {})
    tax_fallback = FALLBACK_TAX_SETTINGS
    tax_override = deep_merge(tax_live, OFFICIAL_2026_OVERRIDES["tax"])
    rows: list[dict[str, Any]] = []
    route = "/settings/impots"

    for slot_ui, year_key, scale_key, ref_year in [
        ("current", "current", "scaleCurrent", 2026),
        ("previous", "previous", "scalePrevious", 2025),
    ]:
        year_label_path = ["incomeTax", "currentYearLabel"] if slot_ui == "current" else ["incomeTax", "previousYearLabel"]
        target_year_value = tax_override["incomeTax"]["currentYearLabel"] if slot_ui == "current" else tax_override["incomeTax"]["previousYearLabel"]
        rows.append(
            build_copy_row(
                route=route,
                section="Bareme de l'impot sur le revenu",
                ui_label="Barreme - libelle d'annee",
                slot_ui=slot_ui,
                annee_reference=ref_year,
                shared_group=f"tax_settings.incomeTax.yearLabel.{slot_ui}",
                supabase_table="tax_settings",
                json_path=stringify_path(year_label_path),
                valeur_live_actuelle=get_path(tax_live, *year_label_path),
                fallback_code=get_path(tax_fallback, *year_label_path),
                valeur_cible=target_year_value,
                unite="texte",
                source_officielle=SOURCE_URLS["millesime"],
                commentaire="Decalage de millesime du plan 2026.",
            )
        )

        for scale_index in range(5):
            for field_key, field_label, unit in [("from", "De", "EUR"), ("to", "A", "EUR"), ("rate", "Taux %", "%")]:
                current_path = ["incomeTax", "scaleCurrent", scale_index, field_key]
                previous_path = ["incomeTax", "scalePrevious", scale_index, field_key]
                (current_live, current_target), (previous_live, previous_target) = build_annual_target(tax_live, tax_fallback, tax_override, current_path, previous_path)
                rows.append(
                    build_copy_row(
                        route=route,
                        section="Bareme de l'impot sur le revenu",
                        ui_label=f"Tranche {scale_index + 1} - {field_label}",
                        slot_ui=slot_ui,
                        annee_reference=ref_year,
                        shared_group=f"tax_settings.incomeTax.scale[{scale_index}].{field_key}",
                        supabase_table="tax_settings",
                        json_path=stringify_path(["incomeTax", scale_key, scale_index, field_key]),
                        valeur_live_actuelle=current_live if slot_ui == "current" else previous_live,
                        fallback_code=get_path(tax_fallback, "incomeTax", scale_key, scale_index, field_key),
                        valeur_cible=current_target if slot_ui == "current" else previous_target,
                        unite=unit,
                        source_officielle=SOURCE_URLS["ir_2026"] if slot_ui == "current" else SOURCE_URLS["baseline_live"],
                        commentaire="Laisser vide pour borne ouverte." if field_key == "to" and (current_target if slot_ui == "current" else previous_target) is None else "",
                    )
                )

        for field_key, field_label in [
            ("plafondPartSup", "Plafond quotient familial - Par 1/2 part supplementaire"),
            ("plafondParentIsoleDeuxPremieresParts", "Plafond quotient familial - Parent isole (2 premieres parts)"),
        ]:
            current_path = ["incomeTax", "quotientFamily", "current", field_key]
            previous_path = ["incomeTax", "quotientFamily", "previous", field_key]
            (current_live, current_target), (previous_live, previous_target) = build_annual_target(tax_live, tax_fallback, tax_override, current_path, previous_path)
            rows.append(
                build_copy_row(
                    route=route,
                    section="Bareme de l'impot sur le revenu",
                    ui_label=field_label,
                    slot_ui=slot_ui,
                    annee_reference=ref_year,
                    shared_group=f"tax_settings.incomeTax.quotientFamily.{field_key}",
                    supabase_table="tax_settings",
                    json_path=stringify_path(["incomeTax", "quotientFamily", year_key, field_key]),
                    valeur_live_actuelle=current_live if slot_ui == "current" else previous_live,
                    fallback_code=get_path(tax_fallback, "incomeTax", "quotientFamily", year_key, field_key),
                    valeur_cible=current_target if slot_ui == "current" else previous_target,
                    unite="EUR",
                    source_officielle=SOURCE_URLS["ir_reductions"] if slot_ui == "current" else SOURCE_URLS["baseline_live"],
                )
            )

        for field_key, field_label, unit in [
            ("triggerSingle", "Decote - Declenchement celibataire", "EUR"),
            ("triggerCouple", "Decote - Declenchement couple", "EUR"),
            ("amountSingle", "Decote - Montant celibataire", "EUR"),
            ("amountCouple", "Decote - Montant couple", "EUR"),
            ("ratePercent", "Decote - Taux de la decote", "%"),
        ]:
            current_path = ["incomeTax", "decote", "current", field_key]
            previous_path = ["incomeTax", "decote", "previous", field_key]
            (current_live, current_target), (previous_live, previous_target) = build_annual_target(tax_live, tax_fallback, tax_override, current_path, previous_path)
            rows.append(
                build_copy_row(
                    route=route,
                    section="Bareme de l'impot sur le revenu",
                    ui_label=field_label,
                    slot_ui=slot_ui,
                    annee_reference=ref_year,
                    shared_group=f"tax_settings.incomeTax.decote.{field_key}",
                    supabase_table="tax_settings",
                    json_path=stringify_path(["incomeTax", "decote", year_key, field_key]),
                    valeur_live_actuelle=current_live if slot_ui == "current" else previous_live,
                    fallback_code=get_path(tax_fallback, "incomeTax", "decote", year_key, field_key),
                    valeur_cible=current_target if slot_ui == "current" else previous_target,
                    unite=unit,
                    source_officielle=SOURCE_URLS["ir_reductions"] if slot_ui == "current" else SOURCE_URLS["baseline_live"],
                )
            )

        for current_key, previous_key, block_label in [
            ("current", "previous", "Abattement 10 %"),
            ("retireesCurrent", "retireesPrevious", "Abattement 10 % pensions retraite"),
        ]:
            for field_key, suffix in [("plafond", "Plafond"), ("plancher", "Plancher")]:
                current_path = ["incomeTax", "abat10", current_key, field_key]
                previous_path = ["incomeTax", "abat10", previous_key, field_key]
                (current_live, current_target), (previous_live, previous_target) = build_annual_target(tax_live, tax_fallback, tax_override, current_path, previous_path)
                key_for_path = current_key if slot_ui == "current" else previous_key
                rows.append(
                    build_copy_row(
                        route=route,
                        section="Bareme de l'impot sur le revenu",
                        ui_label=f"{block_label} - {suffix}",
                        slot_ui=slot_ui,
                        annee_reference=ref_year,
                        shared_group=f"tax_settings.incomeTax.abat10.{current_key}.{field_key}",
                        supabase_table="tax_settings",
                        json_path=stringify_path(["incomeTax", "abat10", key_for_path, field_key]),
                        valeur_live_actuelle=current_live if slot_ui == "current" else previous_live,
                        fallback_code=get_path(tax_fallback, "incomeTax", "abat10", key_for_path, field_key),
                        valeur_cible=current_target if slot_ui == "current" else previous_target,
                        unite="EUR",
                        source_officielle=SOURCE_URLS["ir_reductions"] if slot_ui == "current" else SOURCE_URLS["baseline_live"],
                    )
                )

    for slot_ui, period_key, ref_year in [("current", "current", 2026), ("previous", "previous", 2025)]:
        for zone_key, zone_label in DOM_ZONES:
            for field_key, field_label, unit in [("ratePercent", "Taux %", "%"), ("cap", "Plafond EUR", "EUR")]:
                current_path = ["incomeTax", "domAbatement", "current", zone_key, field_key]
                previous_path = ["incomeTax", "domAbatement", "previous", zone_key, field_key]
                (current_live, current_target), (previous_live, previous_target) = build_annual_target(tax_live, tax_fallback, tax_override, current_path, previous_path)
                live_value = current_live if slot_ui == "current" else previous_live
                target_value = current_target if slot_ui == "current" else previous_target
                rows.append(
                    build_copy_row(
                        route=route,
                        section="Abattement DOM sur l'IR (bareme)",
                        ui_label=f"{zone_label} - {field_label}",
                        slot_ui=slot_ui,
                        annee_reference=ref_year,
                        shared_group=f"tax_settings.incomeTax.domAbatement.{zone_key}.{field_key}",
                        supabase_table="tax_settings",
                        json_path=stringify_path(["incomeTax", "domAbatement", period_key, zone_key, field_key]),
                        valeur_live_actuelle=live_value,
                        fallback_code=get_path(tax_fallback, "incomeTax", "domAbatement", period_key, zone_key, field_key),
                        valeur_cible=target_value,
                        unite=unit,
                        source_officielle=SOURCE_URLS["ir_reductions"] if slot_ui == "current" else SOURCE_URLS["baseline_live"],
                        commentaire="Aucun changement officiel 2026 identifie dans la recherche actuelle." if values_equal(live_value, target_value) else "",
                    )
                )

    for slot_ui, period_key, ref_year in [("current", "current", 2026), ("previous", "previous", 2025)]:
        for field_key, field_label in [("rateIR", "Part impot sur le revenu"), ("rateSocial", "Prelevements sociaux"), ("rateTotal", "Taux global PFU")]:
            current_path = ["pfu", "current", field_key]
            previous_path = ["pfu", "previous", field_key]
            (current_live, current_target), (previous_live, previous_target) = build_annual_target(tax_live, tax_fallback, tax_override, current_path, previous_path)
            comment = ""
            if slot_ui == "current" and field_key == "rateSocial":
                comment = "Hausse 2026 observee sur le PFU: 18,6 %."
            if slot_ui == "current" and field_key == "rateTotal":
                comment = "PFU 2026 = 31,4 %."
            rows.append(
                build_copy_row(
                    route=route,
                    section="PFU (flat tax)",
                    ui_label=field_label,
                    slot_ui=slot_ui,
                    annee_reference=ref_year,
                    shared_group=f"tax_settings.pfu.{field_key}",
                    supabase_table="tax_settings",
                    json_path=stringify_path(["pfu", period_key, field_key]),
                    valeur_live_actuelle=current_live if slot_ui == "current" else previous_live,
                    fallback_code=get_path(tax_fallback, "pfu", period_key, field_key),
                    valeur_cible=current_target if slot_ui == "current" else previous_target,
                    unite="%",
                    source_officielle=SOURCE_URLS["pfu_2026"] if slot_ui == "current" else SOURCE_URLS["baseline_live"],
                    commentaire=comment,
                )
            )

    for slot_ui, period_key, ref_year in [("current", "current", 2026), ("previous", "previous", 2025)]:
        for group_key, group_label in [("single", "CEHR - personne seule"), ("couple", "CEHR - couple")]:
            for index in range(2):
                current_path = ["cehr", "current", group_key, index, "rate"]
                previous_path = ["cehr", "previous", group_key, index, "rate"]
                (current_live, current_target), (previous_live, previous_target) = build_annual_target(tax_live, tax_fallback, tax_override, current_path, previous_path)
                rows.append(
                    build_copy_row(
                        route=route,
                        section="CEHR / CDHR",
                        ui_label=f"{group_label} - tranche {index + 1} - Taux %",
                        slot_ui=slot_ui,
                        annee_reference=ref_year,
                        shared_group=f"tax_settings.cehr.{group_key}[{index}].rate",
                        supabase_table="tax_settings",
                        json_path=stringify_path(["cehr", period_key, group_key, index, "rate"]),
                        valeur_live_actuelle=current_live if slot_ui == "current" else previous_live,
                        fallback_code=get_path(tax_fallback, "cehr", period_key, group_key, index, "rate"),
                        valeur_cible=current_target if slot_ui == "current" else previous_target,
                        unite="%",
                        source_officielle=SOURCE_URLS["ir_reductions"] if slot_ui == "current" else SOURCE_URLS["baseline_live"],
                    )
                )
        for field_key, field_label, unit in [("minEffectiveRate", "CDHR - Taux effectif minimal", "%"), ("thresholdSingle", "CDHR - Seuil RFR personne seule", "EUR"), ("thresholdCouple", "CDHR - Seuil RFR couple", "EUR")]:
            current_path = ["cdhr", "current", field_key]
            previous_path = ["cdhr", "previous", field_key]
            (current_live, current_target), (previous_live, previous_target) = build_annual_target(tax_live, tax_fallback, tax_override, current_path, previous_path)
            rows.append(
                build_copy_row(
                    route=route,
                    section="CEHR / CDHR",
                    ui_label=field_label,
                    slot_ui=slot_ui,
                    annee_reference=ref_year,
                    shared_group=f"tax_settings.cdhr.{field_key}",
                    supabase_table="tax_settings",
                    json_path=stringify_path(["cdhr", period_key, field_key]),
                    valeur_live_actuelle=current_live if slot_ui == "current" else previous_live,
                    fallback_code=get_path(tax_fallback, "cdhr", period_key, field_key),
                    valeur_cible=current_target if slot_ui == "current" else previous_target,
                    unite=unit,
                    source_officielle=SOURCE_URLS["ir_reductions"] if slot_ui == "current" else SOURCE_URLS["baseline_live"],
                )
            )

    for slot_ui, period_key, ref_year in [("current", "current", 2026), ("previous", "previous", 2025)]:
        for field_key, field_label, unit in [("normalRate", "Taux normal IS", "%"), ("reducedRate", "Taux reduit IS", "%"), ("reducedThreshold", "Seuil de benefice au taux reduit", "EUR")]:
            current_path = ["corporateTax", "current", field_key]
            previous_path = ["corporateTax", "previous", field_key]
            (current_live, current_target), (previous_live, previous_target) = build_annual_target(tax_live, tax_fallback, tax_override, current_path, previous_path)
            rows.append(
                build_copy_row(
                    route=route,
                    section="Impot sur les societes",
                    ui_label=field_label,
                    slot_ui=slot_ui,
                    annee_reference=ref_year,
                    shared_group=f"tax_settings.corporateTax.{field_key}",
                    supabase_table="tax_settings",
                    json_path=stringify_path(["corporateTax", period_key, field_key]),
                    valeur_live_actuelle=current_live if slot_ui == "current" else previous_live,
                    fallback_code=get_path(tax_fallback, "corporateTax", period_key, field_key),
                    valeur_cible=current_target if slot_ui == "current" else previous_target,
                    unite=unit,
                    source_officielle=SOURCE_URLS["baseline_live"],
                    commentaire="Aucun changement officiel 2026 trouve dans la recherche actuelle." if slot_ui == "current" else "",
                )
            )

    append_dmtg_rows(
        rows,
        route=route,
        tax_live=tax_live,
        tax_fallback=tax_fallback,
        section="Droits de Mutation a Titre Gratuit (DMTG)",
    )
    return rows


def build_prelevements_rows(snapshot: dict[str, Any]) -> list[dict[str, Any]]:
    ps_live = deep_merge(FALLBACK_PS_SETTINGS, snapshot["ps_settings"].get("data") or {})
    ps_fallback = FALLBACK_PS_SETTINGS
    ps_override = deep_merge(ps_live, OFFICIAL_2026_OVERRIDES["ps"])
    pass_live = build_live_map(snapshot["pass_history"].get("data") or [])
    rows: list[dict[str, Any]] = []
    route = "/settings/prelevements"

    for year in range(2019, 2027):
        live_value = pass_live.get(year, FALLBACK_PASS_HISTORY.get(year))
        fallback_value = FALLBACK_PASS_HISTORY.get(year)
        target_value = OFFICIAL_2026_OVERRIDES["pass_history"].get(year, live_value if live_value is not None else fallback_value)
        rows.append(
            build_copy_row(
                route=route,
                section="Historique du PASS (8 valeurs)",
                ui_label=f"PASS {year}",
                slot_ui=f"pass_{year}",
                annee_reference=year,
                shared_group=f"public.pass_history.{year}",
                supabase_table="pass_history",
                json_path=f"year={year}",
                valeur_live_actuelle=live_value,
                fallback_code=fallback_value,
                valeur_cible=target_value,
                unite="EUR",
                source_officielle=SOURCE_URLS["pass_2026"] if year == 2026 else SOURCE_URLS["baseline_live"],
                commentaire="Le fallback code s'arrete a 2025." if year == 2026 and fallback_value is None else "",
            )
        )

    for slot_ui, period_key, ref_year in [("current", "current", 2026), ("previous", "previous", 2025)]:
        for field_key, label in [("totalRate", "Taux global des prelevements sociaux"), ("csgDeductibleRate", "CSG deductible (au bareme)")]:
            current_path = ["patrimony", "current", field_key]
            previous_path = ["patrimony", "previous", field_key]
            (current_live, current_target), (previous_live, previous_target) = build_annual_target(ps_live, ps_fallback, ps_override, current_path, previous_path)
            rows.append(
                build_copy_row(
                    route=route,
                    section="Prelevements sociaux - patrimoine et capital",
                    ui_label=label,
                    slot_ui=slot_ui,
                    annee_reference=ref_year,
                    shared_group=f"ps_settings.patrimony.{field_key}",
                    supabase_table="ps_settings",
                    json_path=stringify_path(["patrimony", period_key, field_key]),
                    valeur_live_actuelle=current_live if slot_ui == "current" else previous_live,
                    fallback_code=get_path(ps_fallback, "patrimony", period_key, field_key),
                    valeur_cible=current_target if slot_ui == "current" else previous_target,
                    unite="%",
                    source_officielle=SOURCE_URLS["pfu_2026"] if slot_ui == "current" else SOURCE_URLS["baseline_live"],
                    commentaire="SER1 modelise un taux unique; exceptions officielles 2026 non representees." if slot_ui == "current" and field_key == "totalRate" else "",
                )
            )

    for slot_ui, period_key, ref_year in [("current", "current", 2026), ("previous", "previous", 2025)]:
        for bracket_index, bracket_label in enumerate(RETIREMENT_LABELS):
            for field_key, field_label, unit in RETIREMENT_FIELDS:
                current_path = ["retirement", "current", "brackets", bracket_index, field_key]
                previous_path = ["retirement", "previous", "brackets", bracket_index, field_key]
                (current_live, current_target), (previous_live, previous_target) = build_annual_target(ps_live, ps_fallback, ps_override, current_path, previous_path)
                metier_validation = slot_ui == "current" and bracket_index in (2, 3) and field_key in ("maladieRate", "totalRate")
                rows.append(
                    build_copy_row(
                        route=route,
                        section="Prelevements sociaux - pensions de retraite",
                        ui_label=f"{bracket_label} - {field_label}",
                        slot_ui=slot_ui,
                        annee_reference=ref_year,
                        shared_group=f"ps_settings.retirement.brackets[{bracket_index}].{field_key}",
                        supabase_table="ps_settings",
                        json_path=stringify_path(["retirement", period_key, "brackets", bracket_index, field_key]),
                        valeur_live_actuelle=current_live if slot_ui == "current" else previous_live,
                        fallback_code=get_path(ps_fallback, "retirement", period_key, "brackets", bracket_index, field_key),
                        valeur_cible=current_target if slot_ui == "current" else previous_target,
                        unite=unit,
                        source_officielle=SOURCE_URLS["ps_retraite_2026"] if slot_ui == "current" else SOURCE_URLS["baseline_live"],
                        commentaire="F2971 affiche des totaux 7,4 % / 9,1 % hors maladie; conserver la modelisation SER1 et valider metier." if metier_validation else "",
                        metier_validation=metier_validation,
                    )
                )

    for slot_ui, period_key, ref_year in [("current", "current", 2026), ("previous", "previous", 2025)]:
        for region_key, region_label in THRESHOLD_REGIONS:
            for field_key, field_label in THRESHOLD_FIELDS:
                current_path = ["retirementThresholds", "current", region_key, field_key]
                previous_path = ["retirementThresholds", "previous", region_key, field_key]
                (current_live, current_target), (previous_live, previous_target) = build_annual_target(ps_live, ps_fallback, ps_override, current_path, previous_path)
                source_url = SOURCE_URLS["ps_retraite_2026"] if region_key == "metropole" and slot_ui == "current" else SOURCE_URLS["ps_dom_2026"] if slot_ui == "current" else SOURCE_URLS["baseline_live"]
                rows.append(
                    build_copy_row(
                        route=route,
                        section="Seuils de revenus pour la CSG, la CRDS et la CASA (RFR)",
                        ui_label=f"{region_label} - {field_label}",
                        slot_ui=slot_ui,
                        annee_reference=ref_year,
                        shared_group=f"ps_settings.retirementThresholds.{region_key}.{field_key}",
                        supabase_table="ps_settings",
                        json_path=stringify_path(["retirementThresholds", period_key, region_key, field_key]),
                        valeur_live_actuelle=current_live if slot_ui == "current" else previous_live,
                        fallback_code=get_path(ps_fallback, "retirementThresholds", period_key, region_key, field_key),
                        valeur_cible=current_target if slot_ui == "current" else previous_target,
                        unite="EUR",
                        source_officielle=source_url,
                    )
                )

    return rows


def build_succession_rows(snapshot: dict[str, Any]) -> list[dict[str, Any]]:
    tax_live = deep_merge(FALLBACK_TAX_SETTINGS, snapshot["tax_settings"].get("data") or {})
    tax_fallback = FALLBACK_TAX_SETTINGS
    fiscality_live = deep_merge(FALLBACK_FISCALITY_SETTINGS, snapshot["fiscality_settings"].get("data") or {})
    fiscality_fallback = FALLBACK_FISCALITY_SETTINGS
    fiscality_target = deep_merge(fiscality_live, OFFICIAL_2026_OVERRIDES["fiscality"])
    rows: list[dict[str, Any]] = []
    route = "/settings/dmtg-succession"

    append_dmtg_rows(
        rows,
        route=route,
        tax_live=tax_live,
        tax_fallback=tax_fallback,
        section="Droits de Mutation a Titre Gratuit (DMTG)",
    )

    donation_live = get_path(tax_live, "donation") or {}
    donation_fallback = get_path(tax_fallback, "donation") or {}
    for path, label, unit, source_url, comment in [
        (["donation", "rappelFiscalAnnees"], "Duree du rappel fiscal", "ans", SOURCE_URLS["donation"], "Le snapshot live tax_settings.donation est vide; reprendre le fallback valide."),
        (["donation", "donFamilial790G", "montant"], "Montant exonere", "EUR", SOURCE_URLS["donation_790g"], "Le snapshot live tax_settings.donation est vide; reprendre le fallback valide."),
        (["donation", "donFamilial790G", "conditions"], "Conditions", "texte", SOURCE_URLS["donation_790g"], "Le snapshot live tax_settings.donation est vide; reprendre le fallback valide."),
        (["donation", "donManuel", "abattementRenouvellement"], "Renouvellement abattement tous les", "ans", SOURCE_URLS["donation"], "Le snapshot live tax_settings.donation est vide; reprendre le fallback valide."),
    ]:
        live_value = get_path(tax_live, *path)
        fallback_value = get_path(tax_fallback, *path)
        target_value = fallback_value
        rows.append(
            build_copy_row(
                route=route,
                section="Donation & rappel fiscal",
                ui_label=label,
                slot_ui="unique",
                annee_reference=2026,
                shared_group=f"tax_settings.{stringify_path(path)}",
                supabase_table="tax_settings",
                json_path=stringify_path(path),
                valeur_live_actuelle=live_value,
                fallback_code=fallback_value,
                valeur_cible=target_value,
                unite=unit,
                source_officielle=source_url,
                commentaire=comment,
            )
        )

    av_specs = [
        (["assuranceVie", "deces", "agePivotPrimes"], "Age pivot primes (avant/apres)", "ans", SOURCE_URLS["av_deces"], ""),
        (["assuranceVie", "deces", "primesApres1998", "allowancePerBeneficiary"], "Abattement par beneficiaire", "EUR", SOURCE_URLS["av_deces"], ""),
        (["assuranceVie", "deces", "primesApres1998", "brackets", 0, "upTo"], "Bareme par beneficiaire - tranche 1 - Jusqu'a (EUR cumule)", "EUR", SOURCE_URLS["av_deces"], "UI stocke un plafond cumule 852 500 = 152 500 + 700 000."),
        (["assuranceVie", "deces", "primesApres1998", "brackets", 0, "ratePercent"], "Bareme par beneficiaire - tranche 1 - Taux %", "%", SOURCE_URLS["av_deces"], ""),
        (["assuranceVie", "deces", "primesApres1998", "brackets", 1, "upTo"], "Bareme par beneficiaire - tranche 2 - Jusqu'a (EUR cumule)", "EUR", SOURCE_URLS["av_deces"], "Laisser vide pour la tranche ouverte."),
        (["assuranceVie", "deces", "primesApres1998", "brackets", 1, "ratePercent"], "Bareme par beneficiaire - tranche 2 - Taux %", "%", SOURCE_URLS["av_deces"], ""),
        (["assuranceVie", "deces", "apres70ans", "globalAllowance"], "Abattement global (tous beneficiaires)", "EUR", SOURCE_URLS["av_deces"], ""),
    ]
    for path, label, unit, source_url, comment in av_specs:
        rows.append(
            build_copy_row(
                route=route,
                section="Assurance-vie deces (990 I / 757 B)",
                ui_label=label,
                slot_ui="unique",
                annee_reference=2026,
                shared_group=f"fiscality_settings.{stringify_path(path)}",
                supabase_table="fiscality_settings",
                json_path=stringify_path(path),
                valeur_live_actuelle=get_path(fiscality_live, *path),
                fallback_code=get_path(fiscality_fallback, *path),
                valeur_cible=get_path(fiscality_target, *path),
                unite=unit,
                source_officielle=source_url,
                commentaire=comment,
            )
        )

    return rows


def build_inventory_rows(copy_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    inventory: list[dict[str, Any]] = []
    for row in copy_rows:
        path = str(row["json_path"])
        if path.startswith("incomeTax."):
            consumer = GROUP_CONSUMERS["domAbatement"] if ".domAbatement." in path else GROUP_CONSUMERS["incomeTax"]
        elif path.startswith("pfu."):
            consumer = GROUP_CONSUMERS["pfu"]
        elif path.startswith("cehr.") or path.startswith("cdhr."):
            consumer = GROUP_CONSUMERS["cehr_cdhr"]
        elif path.startswith("corporateTax."):
            consumer = GROUP_CONSUMERS["corporateTax"]
        elif path.startswith("dmtg."):
            consumer = GROUP_CONSUMERS["dmtg"]
        elif path.startswith("donation."):
            consumer = GROUP_CONSUMERS["donation"]
        elif path.startswith("assuranceVie.deces."):
            consumer = GROUP_CONSUMERS["av_deces"]
        elif path.startswith("patrimony."):
            consumer = GROUP_CONSUMERS["patrimony"]
        elif path.startswith("retirement.brackets."):
            consumer = GROUP_CONSUMERS["retirement"]
        elif path.startswith("retirementThresholds."):
            consumer = GROUP_CONSUMERS["retirementThresholds"]
        elif str(row["supabase_table"]) == "pass_history":
            consumer = GROUP_CONSUMERS["pass_history"]
        else:
            consumer = {"status": "a_documenter", "proof": ""}

        inventory.append(
            {
                "route": row["route"],
                "section": row["section"],
                "type": "input",
                "shared_group": row["shared_group"],
                "supabase_table": row["supabase_table"],
                "json_path": row["json_path"],
                "editable": "oui",
                "consumer_status": consumer["status"],
                "consumer_proof": consumer["proof"],
                "notes": row["commentaire"],
            }
        )

    inventory.extend(
        [
            {
                "route": "/settings/prelevements",
                "section": "Libelles de millesime",
                "type": "meta_hors_ui",
                "shared_group": "ps_settings.labels.currentYearLabel",
                "supabase_table": "ps_settings",
                "json_path": "labels.currentYearLabel",
                "editable": "non",
                "consumer_status": GROUP_CONSUMERS["ps_labels"]["status"],
                "consumer_proof": GROUP_CONSUMERS["ps_labels"]["proof"],
                "notes": "Visible dans les en-tetes des colonnes settings, non editable depuis la page.",
            },
            {
                "route": "/settings/prelevements",
                "section": "Libelles de millesime",
                "type": "meta_hors_ui",
                "shared_group": "ps_settings.labels.previousYearLabel",
                "supabase_table": "ps_settings",
                "json_path": "labels.previousYearLabel",
                "editable": "non",
                "consumer_status": GROUP_CONSUMERS["ps_labels"]["status"],
                "consumer_proof": GROUP_CONSUMERS["ps_labels"]["proof"],
                "notes": "Visible dans les en-tetes des colonnes settings, non editable depuis la page.",
            },
        ]
    )

    for route, section in READ_ONLY_SECTIONS:
        inventory.append(
            {
                "route": route,
                "section": section,
                "type": "lecture_seule",
                "shared_group": section,
                "supabase_table": "",
                "json_path": "",
                "editable": "non",
                "consumer_status": "documentation metier",
                "consumer_proof": (
                    "ReserveCivilSection.tsx, RegimesSection.tsx, "
                    "LiberalitesSection.tsx, AvantagesMatrimoniauxSection.tsx"
                ),
                "notes": "Section informative non parametree, a conserver dans le dossier d'audit.",
            }
        )

    return inventory


def autosize(ws: Any, widths: dict[int, int] | None = None) -> None:
    for index, column in enumerate(ws.iter_cols(min_row=1, max_row=1), start=1):
        header = str(column[0].value or "")
        width = max(12, min(48, len(header) + 4))
        if widths and index in widths:
            width = widths[index]
        ws.column_dimensions[get_column_letter(index)].width = width


def write_sheet(ws: Any, columns: list[str], rows: list[dict[str, Any]], status_col: str | None = None) -> None:
    ws.append(columns)
    for row in rows:
        ws.append([row.get(column, "") for column in columns])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    if status_col:
        status_index = columns.index(status_col) + 1
        for row_index in range(2, ws.max_row + 1):
            with_status_fill(ws.cell(row=row_index, column=status_index), str(ws.cell(row=row_index, column=status_index).value))
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    autosize(ws)


def assert_counts(copy_rows: list[dict[str, Any]]) -> None:
    counts = {
        "/settings/impots": 125,
        "/settings/prelevements": 112,
        "/settings/dmtg-succession": 48,
    }
    actual: dict[str, int] = {}
    for row in copy_rows:
        actual[row["route"]] = actual.get(row["route"], 0) + 1
    if actual != counts:
        raise RuntimeError(f"Unexpected row counts: {actual} != {counts}")


def main() -> None:
    snapshot = load_snapshot()
    copy_rows = build_impots_rows(snapshot) + build_prelevements_rows(snapshot) + build_succession_rows(snapshot)
    assert_counts(copy_rows)
    inventory_rows = build_inventory_rows(copy_rows)
    source_rows = build_source_rows()
    usage_rows = build_dmtg_usage_rows()

    wb = Workbook()
    ws_copy = wb.active
    ws_copy.title = "01_A_Copier"
    write_sheet(ws_copy, COPY_COLUMNS, copy_rows, status_col="statut")

    ws_inventory = wb.create_sheet("02_Inventaire_Complet")
    write_sheet(ws_inventory, INVENTORY_COLUMNS, inventory_rows)

    ws_sources = wb.create_sheet("03_Sources")
    write_sheet(ws_sources, SOURCE_COLUMNS, source_rows)

    ws_usage = wb.create_sheet("04_Usage_DMTG")
    write_sheet(ws_usage, USAGE_COLUMNS, usage_rows)

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(json.dumps({
        "output": str(OUTPUT_PATH),
        "copy_rows": len(copy_rows),
        "inventory_rows": len(inventory_rows),
        "source_rows": len(source_rows),
        "usage_rows": len(usage_rows),
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
